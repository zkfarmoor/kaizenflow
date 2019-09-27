#!/usr/bin/env python
"""
Download product slates and contract specs from CME.

In more detail: download product slate, save it to xls and xlsx,
_extract hyperlinks from this table, download contract specs from
each of those links and save a dataframe with product slate and
contract specs for each product.

Usage example:
> python vendors/cme/utils.py \
 --download_url https://www.cmegroup.com/CmeWS/mvc/ProductSlate/V1/Download.xls \
 --product_slate_xls_dst_path product_slate.xls \
 --slate_with_specs_csv_dst_path slate_with_specs.csv
"""

# TODO(gp): Dummy for RB.

import argparse
import logging
import os
import string
import time

import numpy as np
import openpyxl
import pandas as pd
import requests
import xlrd
from bs4 import BeautifulSoup
from tqdm.autonotebook import tqdm

import helpers.dbg as dbg

_LOG = logging.getLogger(__name__)

_DOWNLOAD_URL = "https://www.cmegroup.com/CmeWS/mvc/ProductSlate/V1/Download.xls"


class _ProductListDownloader:
    """
    Download product slate as xlsx. CME product slates is a table where for
    each product there is such information as Exchange, Product Group,
    Subgroup, Category, Volume, Open Interest, etc.
    Link to a pruduct slate at the CME website:
    https://www.cmegroup.com/trading/products/?redirect=/trading/index.html#pageNumber=1&sortAsc=false&cleared=Futures

    :param download_url: The download url of the product slate
    :param xls_path: The path to save the product slate xls
    """

    def __init__(self, download_url, xls_path):
        self.download_url = download_url
        dbg.dassert_eq(
            os.path.splitext(self.download_url)[1],
            os.path.splitext(xls_path)[1],
            "Extensions of the files do not match.",
        )
        self.xls_path = xls_path
        self.xlsx_path = os.path.splitext(self.xls_path)[0] + ".xlsx"

    def execute(self, first_row=4):
        """
        Download xls and save the table to xlsx starting with the
        first_row row.

        :param first_row: From which row to save the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1).
        """
        self._download_xls()
        self._save_xls_to_xlsx(first_row=first_row)

    def _download_xls(self):
        """
        Download product slate xls by url.
        """
        response = requests.get(self.download_url)
        if response.status_code == 200:
            with open(self.xls_path, "wb") as f:
                f.write(response.content)
            _LOG.info(f"Downloaded {self.download_url} to {self.xls_path}.")
        else:
            raise ValueError(f"Request status code is {response.status_code}.")

    def _save_xls_to_xlsx(self, first_row):
        """
        Open xls file, convert it to xlsx and save. You can choose from
        which row to save the file.

        :param first_row: From which row to open the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1).
        """
        openpyxl_wb = self._convert_xls_to_openpyxl(first_row=first_row)
        openpyxl_wb.save(self.xlsx_path)
        _LOG.info(f"Converted {self.xls_path} to {self.xlsx_path}.")

    def _convert_xls_to_openpyxl(self, first_row=1):
        """
        Open xls file into openpyxl (openpyxl does not support xls).
        You can choose from which row to save the file.

        :param first_row: From which row to open the file.
            If first_row==1, no rows are dropped (in openpyxl
            the numeration starts at 1)
        :return: openpyxl workbook
        """

        def _excel_loc(row, col):
            letters = string.ascii_uppercase
            result = []
            while col:
                col, rem = divmod(col - 1, 26)
                result[:0] = letters[rem]
            return "".join(result) + str(row)

<<<<<<< HEAD
=======
        # Convert an xls file into openpyxl by copying all the cells
        # from one to the other.
>>>>>>> 39a9e33... PartTask274 lint
        xlrd_book = xlrd.open_workbook(self.xls_path)
        xlrd_sheet = xlrd_book.sheet_by_index(0)
        openpyxl_wb = openpyxl.workbook.Workbook()
        openpyxl_sheet = openpyxl_wb.active
        for i, row in enumerate(range(first_row, xlrd_sheet.nrows)):
            for col in range(0, xlrd_sheet.ncols):
                value = xlrd_sheet.cell_value(row, col)
                openpyxl_sheet.cell(row=i + 1, column=col + 1).value = value
        hyperlink_map = xlrd_sheet.hyperlink_map
        for hyperlink_idx in hyperlink_map.keys():
            row_id = hyperlink_idx[0] - first_row + 1
            col_id = hyperlink_idx[1] + 1
            hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=_excel_loc(row_id, col_id),
                display=openpyxl_sheet.cell(row=row_id, column=col_id).value,
                target=hyperlink_map[hyperlink_idx].url_or_path,
            )
            openpyxl_sheet.cell(row=row_id, column=col_id).hyperlink = hyperlink
        return openpyxl_wb


class _HyperlinksExtractor:
    """
    Extract hyperlinks from a table and save it with a hyperlinks
    column.

    :param xlsx_path: a path to the xlsx table
    """

    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.hyperlinks = []
        self.df_with_hyperlinks = pd.DataFrame()

    def execute(
        self, dst_path, hyperlink_col_loc=5, hyperlink_col_name="product_link"
    ):
        """
        Extract hyperlinks from the dataframe and save a dataframe with
        a hyperlinks column.

        :param dst_path: a path for saving the resulting dataframe
        :param hyperlink_col_loc: Number of column from which to _extract
            the hyperlinks. In openpyxl numeration starts at 1.
        :param hyperlink_col_name: The name of the hyperlinks column
            in the output dataframe
        """
        self._get_df_with_hyperlinks(
            hyperlink_col_loc=hyperlink_col_loc,
            hyperlink_col_name=hyperlink_col_name,
        )
        self.df_with_hyperlinks.to_excel(dst_path)
        _LOG.info(f"Saved dataframe with hyperlinks to {dst_path}.")

    def _extract(self, hyperlink_col_loc):
        """
        Get hyperlinks for the xlsx file.

        :param hyperlink_col_loc: From which column to _extract hyperlinks.
            In openpyxl the numeration starts at 1.
        :return: A list of hyperlinks.
        """
        workbook = openpyxl.load_workbook(self.xlsx_path)
        ws = workbook[workbook.get_sheet_names()[0]]
        hyperlinks = []
        for row in ws.iter_rows(
            min_row=2, min_col=hyperlink_col_loc, max_col=hyperlink_col_loc
        ):
            hyperlinks.append(row[0].hyperlink.target)
        return hyperlinks

    def _add_hyperlinks_to_df(self, hyperlink_col_name):
        """
        Read xlsx dataframe and add hyperlinks as a column.

        :param hyperlink_col_name: the name of the hyperlinks column
            in the output dataframe
        :return: pd.DataFrame with hyperlinks column
        """
        df = pd.read_excel(self.xlsx_path)
        df[hyperlink_col_name] = self.hyperlinks
        return df

    def _get_df_with_hyperlinks(
        self, hyperlink_col_loc=5, hyperlink_col_name="product_link"
    ):
        """
        Extract hyperlinks from the dataframe and get a dataframe with
        a hyperlinks column.

        :param hyperlink_col_loc: Number of column from which to _extract
            the hyperlinks. In openpyxl numeration starts at 1.
        :param hyperlink_col_name: The name of the hyperlinks column
            in the output dataframe
        :return: A dataframe with a hyperlinks column
        """
        self.hyperlinks = self._extract(hyperlink_col_loc=hyperlink_col_loc)
        self.df_with_hyperlinks = self._add_hyperlinks_to_df(
            hyperlink_col_name=hyperlink_col_name
        )


class HTMLTablesDownloader:
    """
    Download html by link, _extract tables with hyperlinks from it,
    concatenate them into one dataframe.

    :param html_url: html link
    """

    def __init__(self, html_url):
        self.html_url = html_url

    def execute(self):
        """
        Download html by link, _extract tables with hyperlinks from it,
        concatenate them into one dataframe.

        :return: pd.DataFrame of the tables extracted from the html.
            If a column has hyperlinks, they will be extracted into a
            column with a prefix "link_".
        """
        req_res = requests.get(self.html_url)
        if req_res.status_code == 200:
            soup = BeautifulSoup(req_res.content, "lxml")
            soup_tables = soup.find_all("table")
            dfs = [
                self._soup_table_to_df_with_links(soup_table)
                for soup_table in soup_tables
            ]
            if len(dfs) > 0:
                concatenated_df = pd.concat(dfs, sort=False)
            else:
                _LOG.info("No tables were extracted from %s", self.html_url)
                concatenated_df = None
        else:
            _LOG.warning(f"Request status code is {req_res.status_code}")
            concatenated_df = None
        return concatenated_df

    @staticmethod
    def _extract_urls(soup_table):
        """
        Get hyperlinks from each cell of a beautiful soup table.

        :param soup_table: beautiful soup table
        :return: [[url]], the shape is (n_rows, n_cols). If there was no url
            in the cell, the corresponding element will be None
        """
        urls_in_rows = []
        for row in soup_table.find_all("tr"):
            urls_in_cols = []
            for td in row.find_all("td"):
                if td.find("a"):
                    url = td.a["href"]
                else:
                    url = None
                urls_in_cols.append(url)
            urls_in_rows.append(urls_in_cols)
        links_df = pd.DataFrame(urls_in_rows).dropna(axis=1, how="all")
        links_df = links_df.add_prefix("link_")
        return links_df

    @staticmethod
    def _soup_table_to_df(soup_table):
        """
        Read beautiful soup table to pandas DataFrame.

        :param soup_table: beautiful soup table
        :return: pd.DataFrame
        """
        return pd.read_html(str(soup_table).replace("colspan", ""))[0]

    @staticmethod
    def _soup_table_to_df_with_links(soup_table):
        """
        Read beautiful soup table to pandas DataFrame. If a column contains
        hyperlinks, output dataframe will have this column with a prefix
        "link_"and extracted links. Elements that do not have a link
        will be None.

        :param soup_table: beautiful soup table
        :return: pd.DataFrame from that table with columns
            for extracted links.
        """
        df_without_links = HTMLTablesDownloader._soup_table_to_df(soup_table)
        links_df = HTMLTablesDownloader._extract_urls(soup_table)
        df_with_links = df_without_links.join(links_df)
        return df_with_links


class ContractSpecsDownloader:
    """
    Download contract specs and merge them to the product slate.
    save the resulting dataframe.

    :param product_slate: pd.DataFrame with the product slate
    """

    def __init__(self, product_slate):
        self.product_slate = product_slate
        self.name_link = (
            self.product_slate.loc[:, ["Product Name", "product_link"]]
            .rename(columns={"Product Name": "product_name"})
            .set_index("product_name")
        )
        self.names_specs_dict = {}
        self.specs_df = pd.DataFrame()
        self.product_slate_with_specs = pd.DataFrame()

    def execute(self, dst_path):
        """
        Download contract specs and merge them to the product slate.
        save the resulting dataframe.

        :param dst_path: Destination for the product slate with
            contract specs csv
        """
        self._get_contract_specs()
        _LOG.info(f"Saved product slate with contract specs to {dst_path}.")
        self.product_slate_with_specs.to_csv(dst_path)

    def _load_htmls(self):
        """
        Load htmls from links and construct a dictionary with them.

        :return: {product_name: html}
        """
        product_names_htmls = {}
        for name, link in tqdm(
            self.name_link.iterrows(), total=len(self.name_link)
        ):
            time.sleep(1)
            html = HTMLTablesDownloader.execute(link[0])
            if html is not None:
                html.set_index(0, inplace=True)
            product_names_htmls[name] = html
        return product_names_htmls

    def _specs_dict_to_df(self):
        """
        Transform contract specs dictionary to a dataframe.

        :return: pd.DataFrame with product_name as index and contract
            spec html indices as columns
        """
        html_rows = [
            self._get_row(html, product_name)
            if html is not None
            else pd.DataFrame(index=[product_name])
            for product_name, html in self.names_specs_dict.items()
        ]
        html_rows = [self._rename_duplicate_cols(html) for html in html_rows]
        specs_df = pd.concat(html_rows, sort=False)
        return specs_df

    def _get_contract_specs(self):
        """
        Download contract specs and merge them to the product slate.

        :return: pd.DataFrame with product slate and contract specs
            for each product
        """
        self.names_specs_dict = self._load_htmls()
        self.specs_df = self._specs_dict_to_df()
        self.product_slate_with_specs = self.product_slate.merge(
            self.specs_df, left_on="Product Name", right_index=True
        )

    @staticmethod
    def _get_squash_cols(df):
        """
        Get names of all columns with int names that are above zero.

        :param df: pd.DataFrame
        :return: list of str column names
        """
        squash_cols = []
        for col_name in df.columns:
            if isinstance(col_name, int):
                if col_name > 0:
                    squash_cols.append(col_name)
        return squash_cols

    @staticmethod
    def _squash_cols(df, cols):
        if cols == [1]:
            squashed_series = df[1]
        else:
            squashed_series = df[cols].fillna("").apply(" ".join, axis=1)
        return squashed_series

    @staticmethod
    def _add_links_as_rows(df):
        link_df = df[["link_1"]].dropna()
        link_df.columns = [1]
        series = pd.concat([df, link_df.rename("{} Link".format)])[[1]]
        return series

    @staticmethod
    def _get_row(df, idx):
        df = df.copy()
        df[1] = ContractSpecsDownloader._squash_cols(
            df, ContractSpecsDownloader._get_squash_cols(df)
        )
        df = df[[1, "link_1"]]
        tr_df = ContractSpecsDownloader._add_links_as_rows(df).T
        tr_df.index = [idx]
        return tr_df

    @staticmethod
    def _rename_duplicate_cols(df):
        df = df.copy()
        dupe_mask = df.columns.duplicated(keep="first")
        duped_col_names = [
            f"{col_name}_{i}" for i, col_name in enumerate(df.columns[dupe_mask])
        ]
        new_index = np.array(df.columns)
        new_index[dupe_mask] = duped_col_names
        df.columns = new_index
        return df


def _get_slate_with_specs_pipeline(
    download_url, product_slate_xls_dst_path, slate_with_specs_csv_dst_path
):
    """
    - Download product list (slate) as xls and xlsx
    - Extract hyperlinks to the contract specs
    - Download contract specs
    - Save a dataframe with the product slate and contract specs for
        each product

    :param download_url: The url from which to download the product
        slate xls.
    :param product_slate_xls_dst_path: The path to save the product
        slate xls.
    :param slate_with_specs_csv_dst_path: The path to save the product
        slate with contract specs csv
    """
    pld = _ProductListDownloader(
        download_url=download_url, xls_path=product_slate_xls_dst_path
    )
    pld.execute()

    xlsx_with_hyperlinks_path = (
        os.path.splitext(pld.xlsx_path)[0] + "_with_hyperlinks.xlsx"
    )
    he = _HyperlinksExtractor(pld.xlsx_path)
    he.execute(xlsx_with_hyperlinks_path)

    product_slate = pd.read_excel(xlsx_with_hyperlinks_path)
    csd = ContractSpecsDownloader(product_slate)
    csd.execute(slate_with_specs_csv_dst_path)


if __name__ == "__main__":
<<<<<<< HEAD
=======
    _DOWNLOAD_URL = (
        "https://www.cmegroup.com/CmeWS/mvc/ProductSlate/V1/Download.xls"
    )
>>>>>>> 39a9e33... PartTask274 lint
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--download_url",
        required=False,
        action="store",
        default=_DOWNLOAD_URL,
        type=str,
    )
<<<<<<< HEAD
    parser.add_argument(
        "--product_slate_xls_dst_path", required=True, action="store", type=str
    )
=======
    parser.add_argument("--product_list", required=True, action="store", type=str)
>>>>>>> 39a9e33... PartTask274 lint
    parser.add_argument(
        "--slate_with_specs_csv_dst_path", required=True, action="store", type=str
    )
    parser.add_argument(
        "-v",
        dest="log_level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Set the logging level",
    )
    args = parser.parse_args()
    dbg.init_logger(args.log_level)
<<<<<<< HEAD
    _get_slate_with_specs_pipeline(
        download_url=args.download_url,
        product_slate_xls_dst_path=args.product_slate_xls_dst_path,
        slate_with_specs_csv_dst_path=args.slate_with_specs_csv_dst_path,
=======
    _get_list_with_specs_pipeline(
        args.download_url, args.product_list, args.product_specs
>>>>>>> 39a9e33... PartTask274 lint
    )
